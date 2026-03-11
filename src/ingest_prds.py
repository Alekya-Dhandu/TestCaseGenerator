from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook

from src.data_loader import load_prd
from src.knowledge_store import INDEX_PATH


PRD_DIR = Path("data/prds")
TESTCASE_DIR = Path("data/testcases")


def _stem(path: Path) -> str:
    return path.stem.lower()


def _load_example_testcases(xlsx_path: Path, max_rows: int = 20) -> List[Dict[str, Any]]:
    wb = load_workbook(filename=str(xlsx_path), read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        return []

    headers = [str(h) if h is not None else "" for h in headers]
    results: List[Dict[str, Any]] = []
    for i, row in enumerate(rows_iter):
        if i >= max_rows:
            break
        rec: Dict[str, Any] = {}
        for col_name, value in zip(headers, row):
            if col_name:
                rec[col_name] = value
        if rec:
            results.append(rec)
    return results


def build_index() -> None:
    PRD_DIR.mkdir(parents=True, exist_ok=True)
    TESTCASE_DIR.mkdir(parents=True, exist_ok=True)

    prd_files = sorted(PRD_DIR.glob("*.pdf"))
    if not prd_files:
        print("No PRD PDFs found in data/prds. Place your PRDs there and rerun.")
        return

    print(f"Found {len(prd_files)} PRD(s). Building knowledge index...")
    documents: List[Dict[str, Any]] = []

    testcase_files = {f.stem.lower(): f for f in TESTCASE_DIR.glob("*.xlsx")}

    for idx, prd_path in enumerate(prd_files, start=1):
        prd_id = f"prd_{idx:03d}"
        stem = _stem(prd_path)
        print(f"- Ingesting {prd_path.name} as {prd_id}")

        text = load_prd(str(prd_path))
        # Very simple heuristic: treat parts of filename (excluding 'prd') as screens/tags.
        parts = [p for p in stem.replace("-", "_").split("_") if p and p not in {"prd"}]
        screens = [p.title() for p in parts]

        example_testcases: List[Dict[str, Any]] = []
        if stem in testcase_files:
            try:
                example_testcases = _load_example_testcases(testcase_files[stem])
                print(f"  Attached {len(example_testcases)} example test case row(s) from {testcase_files[stem].name}")
            except Exception as exc:
                print(f"  Warning: failed to read {testcase_files[stem].name}: {exc}")

        documents.append(
            {
                "id": prd_id,
                "filename": prd_path.name,
                "text": text,
                "screens": screens,
                "example_testcases": example_testcases,
            }
        )

    INDEX_PATH.parent.mkdir(parents=True, exist_ok=True)
    INDEX_PATH.write_text(json.dumps({"documents": documents}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote index with {len(documents)} document(s) to {INDEX_PATH}")


if __name__ == "__main__":
    build_index()

