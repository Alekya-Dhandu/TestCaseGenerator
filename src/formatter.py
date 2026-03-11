from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def _preferred_columns() -> List[str]:
    # Stable, import-friendly column set (good default for TestPad-like tools).
    return [
        "ID",
        "Title",
        "Type",
        "Priority",
        "Preconditions",
        "Steps",
        "ExpectedResult",
        "Screen",
        "Tags",
    ]


def _normalise_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        # For Steps prefer newline; for Tags we will comma-join below.
        return "\n".join(str(v) for v in value)
    return str(value)


def _normalise_row(tc: Dict[str, Any], columns: Sequence[str]) -> List[str]:
    row: List[str] = []
    for col in columns:
        value = tc.get(col, "")
        if col == "Tags" and isinstance(value, (list, tuple)):
            row.append(", ".join(str(v) for v in value))
        else:
            row.append(_normalise_value(value))
    return row


def _compute_columns(test_cases_list: List[Dict[str, Any]]) -> List[str]:
    preferred = _preferred_columns()
    extras: List[str] = []
    for tc in test_cases_list:
        for k in tc.keys():
            if k not in preferred and k not in extras:
                extras.append(k)
    return preferred + extras


def _auto_fit_columns(ws, max_width: int = 60) -> None:
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            val = cell.value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_width)


def format_to_excel(
    test_cases_list: List[Dict[str, Any]],
    output_path: str = "data/outputs/generated_tests.xlsx",
) -> str:
    """
    Existing helper: write an .xlsx file to disk and return the DataFrame.
    """
    excel_bytes = build_excel_bytes(test_cases_list)
    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_bytes(excel_bytes.getvalue())
    print(f"Saved to {output_path}")
    return output_path


def build_excel_bytes(test_cases_list: List[Dict[str, Any]]) -> BytesIO:
    """
    Build an in-memory .xlsx file (for API download) and return a BytesIO.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    columns = _compute_columns(test_cases_list)

    # Header
    header_font = Font(bold=True)
    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=col)
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Rows
    wrap = Alignment(vertical="top", wrap_text=True)
    for r_idx, tc in enumerate(test_cases_list, start=2):
        row = _normalise_row(tc, columns)
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = wrap

    ws.freeze_panes = "A2"
    _auto_fit_columns(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# CSV version for reference:
# df.to_csv('test_cases.csv', index=False)